import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#CORRECTING prices in excel workSheet
#giving it an alias to make code shorter
def process_workbook(excel_filename):
    
    #loading an excel workbook and return a workbook object
    wb = xl.load_workbook(excel_filename)
    #specify name of sheet we are working with which has various cells
    sheet = wb['Sheet1']
    #With sheet object you can give the name of cell using coordinates
    # cell = sheet['a1']
    # cell = sheet.cell(1,1)
    # print(cell.value)

    #Generating number all the way to the maximum row in the sheet
    for row in range(2,sheet.max_row + 1):
        #Each object in each box is a cell object
        #getting all required values in third column
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price
        
    #saving updated excel
    #Using the reference class to reference particular values
    #we only want values in row 2 to 4 and only column 4
    #we are creating an instance of the reference class and storing it in a variable
    values = Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4 )
    #creating an instance of barchart class
    chart = BarChart()
    #Adding the values you want to construct
    chart.add_data(values)
    #add chart to sheet and specify where
    sheet.add_chart(chart,"e2")

    wb.save(excel_filename)
    
process_workbook("transactions.xlsx")
print("Project done!")